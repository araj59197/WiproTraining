import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.dimensions import ColumnDimension

def create_enhanced_locator_mapping():
    """Create a beautifully formatted Excel file with TutorialsNinja registration locators including relative and absolute XPaths"""
    
    # Enhanced locator data with both relative and absolute XPath
    locator_data = [
        {
            'Element Name': 'First Name',
            'Field ID': 'input-firstname',
            'Field Name': 'firstname',
            'Link Text': 'N/A',
            'Partial Link Text': 'N/A',
            'CSS (Tag & ID)': 'input#input-firstname',
            'CSS (Tag & Class)': 'input.form-control',
            'CSS (Tag & Attribute)': "input[placeholder='First Name']",
            'CSS (Tag, Class & Attribute)': "input.form-control[placeholder='First Name']",
            'CSS (Inner Text)': 'N/A',
            'XPath (ID)': "//input[@id='input-firstname']",
            'XPath (Name)': "//input[@name='firstname']",
            'XPath (Placeholder)': "//input[@placeholder='First Name']",
            'Relative XPath': "//div[@class='form-group required']//input[@id='input-firstname']",
            'Absolute XPath': "/html/body/div[2]/div/div/form/div/div/div[2]/div/input[@id='input-firstname']"
        },
        {
            'Element Name': 'Last Name',
            'Field ID': 'input-lastname',
            'Field Name': 'lastname',
            'Link Text': 'N/A',
            'Partial Link Text': 'N/A',
            'CSS (Tag & ID)': 'input#input-lastname',
            'CSS (Tag & Class)': 'input.form-control',
            'CSS (Tag & Attribute)': "input[placeholder='Last Name']",
            'CSS (Tag, Class & Attribute)': "input.form-control[placeholder='Last Name']",
            'CSS (Inner Text)': 'N/A',
            'XPath (ID)': "//input[@id='input-lastname']",
            'XPath (Name)': "//input[@name='lastname']",
            'XPath (Placeholder)': "//input[@placeholder='Last Name']",
            'Relative XPath': "//div[@class='form-group required']//input[@id='input-lastname']",
            'Absolute XPath': "/html/body/div[2]/div/div/form/div/div/div[3]/div/input[@id='input-lastname']"
        },
        {
            'Element Name': 'Email Address',
            'Field ID': 'input-email',
            'Field Name': 'email',
            'Link Text': 'N/A',
            'Partial Link Text': 'N/A',
            'CSS (Tag & ID)': 'input#input-email',
            'CSS (Tag & Class)': 'input.form-control',
            'CSS (Tag & Attribute)': "input[placeholder='E-Mail']",
            'CSS (Tag, Class & Attribute)': "input.form-control[placeholder='E-Mail']",
            'CSS (Inner Text)': 'N/A',
            'XPath (ID)': "//input[@id='input-email']",
            'XPath (Name)': "//input[@name='email']",
            'XPath (Placeholder)': "//input[@placeholder='E-Mail']",
            'Relative XPath': "//div[@class='form-group required']//input[@id='input-email']",
            'Absolute XPath': "/html/body/div[2]/div/div/form/div/div/div[4]/div/input[@id='input-email']"
        },
        {
            'Element Name': 'Telephone Number',
            'Field ID': 'input-telephone',
            'Field Name': 'telephone',
            'Link Text': 'N/A',
            'Partial Link Text': 'N/A',
            'CSS (Tag & ID)': 'input#input-telephone',
            'CSS (Tag & Class)': 'input.form-control',
            'CSS (Tag & Attribute)': "input[placeholder='Telephone']",
            'CSS (Tag, Class & Attribute)': "input.form-control[placeholder='Telephone']",
            'CSS (Inner Text)': 'N/A',
            'XPath (ID)': "//input[@id='input-telephone']",
            'XPath (Name)': "//input[@name='telephone']",
            'XPath (Placeholder)': "//input[@placeholder='Telephone']",
            'Relative XPath': "//div[@class='form-group required']//input[@id='input-telephone']",
            'Absolute XPath': "/html/body/div[2]/div/div/form/div/div/div[5]/div/input[@id='input-telephone']"
        },
        {
            'Element Name': 'Password',
            'Field ID': 'input-password',
            'Field Name': 'password',
            'Link Text': 'N/A',
            'Partial Link Text': 'N/A',
            'CSS (Tag & ID)': 'input#input-password',
            'CSS (Tag & Class)': 'input.form-control',
            'CSS (Tag & Attribute)': "input[type='password'][placeholder='Password']",
            'CSS (Tag, Class & Attribute)': "input.form-control[type='password'][placeholder='Password']",
            'CSS (Inner Text)': 'N/A',
            'XPath (ID)': "//input[@id='input-password']",
            'XPath (Name)': "//input[@name='password']",
            'XPath (Placeholder)': "//input[@placeholder='Password']",
            'Relative XPath': "//fieldset[@id='account']//input[@id='input-password']",
            'Absolute XPath': "/html/body/div[2]/div/div/form/fieldset[2]/div/div/input[@id='input-password']"
        },
        {
            'Element Name': 'Confirm Password',
            'Field ID': 'input-confirm',
            'Field Name': 'confirm',
            'Link Text': 'N/A',
            'Partial Link Text': 'N/A',
            'CSS (Tag & ID)': 'input#input-confirm',
            'CSS (Tag & Class)': 'input.form-control',
            'CSS (Tag & Attribute)': "input[type='password'][placeholder='Password Confirm']",
            'CSS (Tag, Class & Attribute)': "input.form-control[type='password'][placeholder='Password Confirm']",
            'CSS (Inner Text)': 'N/A',
            'XPath (ID)': "//input[@id='input-confirm']",
            'XPath (Name)': "//input[@name='confirm']",
            'XPath (Placeholder)': "//input[@placeholder='Password Confirm']",
            'Relative XPath': "//fieldset[@id='account']//input[@id='input-confirm']",
            'Absolute XPath': "/html/body/div[2]/div/div/form/fieldset[2]/div/div[2]/input[@id='input-confirm']"
        },
        {
            'Element Name': 'Newsletter Subscribe - Yes',
            'Field ID': 'N/A',
            'Field Name': 'newsletter',
            'Link Text': 'N/A',
            'Partial Link Text': 'N/A',
            'CSS (Tag & ID)': 'N/A',
            'CSS (Tag & Class)': 'N/A',
            'CSS (Tag & Attribute)': "input[type='radio'][value='1'][name='newsletter']",
            'CSS (Tag, Class & Attribute)': "input[type='radio'][value='1'][name='newsletter']",
            'CSS (Inner Text)': 'N/A',
            'XPath (ID)': 'N/A',
            'XPath (Name)': "//input[@name='newsletter'][@value='1']",
            'XPath (Placeholder)': "//input[@type='radio'][@name='newsletter'][@value='1']",
            'Relative XPath': "//fieldset//div[@class='form-group']//input[@name='newsletter'][@value='1']",
            'Absolute XPath': "/html/body/div[2]/div/div/form/fieldset[3]/div/div/div/label/input[@name='newsletter'][@value='1']"
        },
        {
            'Element Name': 'Newsletter Subscribe - No',
            'Field ID': 'N/A',
            'Field Name': 'newsletter',
            'Link Text': 'N/A',
            'Partial Link Text': 'N/A',
            'CSS (Tag & ID)': 'N/A',
            'CSS (Tag & Class)': 'N/A',
            'CSS (Tag & Attribute)': "input[type='radio'][value='0'][name='newsletter']",
            'CSS (Tag, Class & Attribute)': "input[type='radio'][value='0'][name='newsletter']",
            'CSS (Inner Text)': 'N/A',
            'XPath (ID)': 'N/A',
            'XPath (Name)': "//input[@name='newsletter'][@value='0']",
            'XPath (Placeholder)': "//input[@type='radio'][@name='newsletter'][@value='0']",
            'Relative XPath': "//fieldset//div[@class='form-group']//input[@name='newsletter'][@value='0']",
            'Absolute XPath': "/html/body/div[2]/div/div/form/fieldset[3]/div/div/div[2]/label/input[@name='newsletter'][@value='0']"
        },
        {
            'Element Name': 'Privacy Policy Checkbox',
            'Field ID': 'N/A',
            'Field Name': 'agree',
            'Link Text': 'N/A',
            'Partial Link Text': 'N/A',
            'CSS (Tag & ID)': 'N/A',
            'CSS (Tag & Class)': 'N/A',
            'CSS (Tag & Attribute)': "input[type='checkbox'][name='agree']",
            'CSS (Tag, Class & Attribute)': "input[type='checkbox'][name='agree']",
            'CSS (Inner Text)': 'N/A',
            'XPath (ID)': 'N/A',
            'XPath (Name)': "//input[@name='agree']",
            'XPath (Placeholder)': "//input[@type='checkbox'][@name='agree']",
            'Relative XPath': "//div[@class='buttons']//input[@name='agree']",
            'Absolute XPath': "/html/body/div[2]/div/div/form/div/div/input[@name='agree']"
        },
        {
            'Element Name': 'Continue/Submit Button',
            'Field ID': 'N/A',
            'Field Name': 'N/A',
            'Link Text': 'N/A',
            'Partial Link Text': 'N/A',
            'CSS (Tag & ID)': 'N/A',
            'CSS (Tag & Class)': 'input.btn.btn-primary',
            'CSS (Tag & Attribute)': "input[type='submit'][value='Continue']",
            'CSS (Tag, Class & Attribute)': "input.btn.btn-primary[type='submit'][value='Continue']",
            'CSS (Inner Text)': "input:contains('Continue')",
            'XPath (ID)': 'N/A',
            'XPath (Name)': "//input[@type='submit'][@value='Continue']",
            'XPath (Placeholder)': "//input[@type='submit' and @value='Continue']",
            'Relative XPath': "//div[@class='buttons']//input[@type='submit'][@value='Continue']",
            'Absolute XPath': "/html/body/div[2]/div/div/form/div/div[2]/input[@type='submit'][@value='Continue']"
        }
    ]
    
    return locator_data

def apply_beautiful_formatting(workbook, worksheet, df):
    """Apply beautiful formatting to the Excel worksheet"""
    
    # Define colors and styles
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    
    # Alternating row colors
    light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Special colors for XPath columns
    xpath_rel_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")  # Light blue for relative XPath
    xpath_abs_fill = PatternFill(start_color="FFF2E6", end_color="FFF2E6", fill_type="solid")  # Light orange for absolute XPath
    
    # Border style
    thin_border = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin")
    )
    
    # Apply header formatting
    for col in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=3, column=col)  # Row 3 because we have title and spacing
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Apply data formatting
    for row in range(4, len(df) + 4):
        # Alternate row colors
        fill_color = light_fill if row % 2 == 0 else white_fill
        
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.border = thin_border
            
            # Special highlighting for XPath columns
            col_name = df.columns[col-1]
            if 'Relative XPath' in col_name:
                cell.fill = xpath_rel_fill
            elif 'Absolute XPath' in col_name:
                cell.fill = xpath_abs_fill
            else:
                cell.fill = fill_color
                
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            # Special formatting for N/A values
            if cell.value == "N/A":
                cell.font = Font(color="999999", italic=True)
    
    # Set column widths for better readability (updated for new columns)
    column_widths = {
        'A': 25,  # Element Name
        'B': 18,  # Field ID
        'C': 15,  # Field Name
        'D': 12,  # Link Text
        'E': 18,  # Partial Link Text
        'F': 25,  # CSS (Tag & ID)
        'G': 20,  # CSS (Tag & Class)
        'H': 35,  # CSS (Tag & Attribute)
        'I': 45,  # CSS (Tag, Class & Attribute)
        'J': 15,  # CSS (Inner Text)
        'K': 30,  # XPath (ID)
        'L': 25,  # XPath (Name)
        'M': 35,  # XPath (Placeholder)
        'N': 45,  # Relative XPath
        'O': 65   # Absolute XPath
    }
    
    for col, width in column_widths.items():
        worksheet.column_dimensions[col].width = width
    
    # Set row height for header
    worksheet.row_dimensions[3].height = 30
    
    # Set row heights for data rows
    for row in range(4, len(df) + 4):
        worksheet.row_dimensions[row].height = 25

def main():
    print("🎨 Creating Enhanced TutorialsNinja Registration Locator Mapping with Relative & Absolute XPaths...")
    
    # Create the enhanced locator data
    data = create_enhanced_locator_mapping()
    
    # Create DataFrame
    df = pd.DataFrame(data)
    
    # Create a new workbook with openpyxl for better formatting control
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Registration Form Locators"
    
    # Add title row at the top
    title_cell = worksheet.cell(row=1, column=1)
    title_cell.value = "TutorialsNinja Registration Form - Enhanced Element Locator Mapping"
    title_cell.font = Font(size=16, bold=True, color="2F5597")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Merge title across all columns
    worksheet.merge_cells(f'A1:{chr(64 + len(df.columns))}1')
    worksheet.row_dimensions[1].height = 40
    
    # Add some spacing
    worksheet.row_dimensions[2].height = 10
    
    # Add data to worksheet starting from row 3
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 3):
        for c_idx, value in enumerate(row, 1):
            worksheet.cell(row=r_idx, column=c_idx, value=value)
    
    # Apply beautiful formatting
    apply_beautiful_formatting(workbook, worksheet, df)
    
    # Save the enhanced Excel file
    excel_file = "TutorialsNinja_Registration_Locators_Beautiful.xlsx"
    try:
        workbook.save(excel_file)
        print(f"✅ Enhanced Excel file created: {excel_file}")
    except Exception as e:
        print(f"❌ Error creating Excel file: {e}")
        return
    
    # Also create a clean CSV version
    csv_file = "TutorialsNinja_Registration_Locators_Enhanced.csv"
    try:
        df.to_csv(csv_file, index=False)
        print(f"✅ Enhanced CSV file created: {csv_file}")
    except Exception as e:
        print(f"❌ Error creating CSV file: {e}")
    
    # Print summary
    print(f"\n📊 Summary:")
    print(f"✨ Total elements mapped: {len(data)}")
    print(f"📋 Locator types included per element:")
    print(f"   • Element ID and Name")
    print(f"   • CSS Selectors (4 variations)")
    print(f"   • XPath Selectors (3 basic + 2 enhanced)")
    print(f"   • Relative XPath (context-based)")
    print(f"   • Absolute XPath (full DOM path)")
    print(f"   • Link Text locators where applicable")
    print(f"\n🎨 Enhanced Formatting features:")
    print(f"   • Professional header with blue background")
    print(f"   • Alternating row colors for better readability")
    print(f"   • Special highlighting for Relative XPath (light blue)")
    print(f"   • Special highlighting for Absolute XPath (light orange)")
    print(f"   • Optimized column widths for new columns")
    print(f"   • Proper cell alignment and text wrapping")
    print(f"   • Special styling for N/A values")
    print(f"   • Border styling throughout")
    
    print(f"\n🔍 XPath Enhancement Details:")
    print("=" * 100)
    print(f"📍 RELATIVE XPATH:")
    print(f"   • Context-aware paths using parent elements")
    print(f"   • More maintainable when DOM structure changes")
    print(f"   • Examples: //div[@class='form-group required']//input[@id='input-firstname']")
    print(f"")
    print(f"📍 ABSOLUTE XPATH:")
    print(f"   • Complete path from HTML root")
    print(f"   • Precise but fragile if DOM changes")
    print(f"   • Examples: /html/body/div[2]/div/div/form/div/div/div[2]/div/input[@id='input-firstname']")
    print("=" * 100)
    
    # Show enhanced preview
    print(f"\n📋 Enhanced Locator Mapping Preview:")
    for i, item in enumerate(data[:2], 1):
        print(f"{i}. {item['Element Name']}:")
        print(f"   ID: {item['Field ID']}")
        print(f"   CSS: {item['CSS (Tag & ID)']}")
        print(f"   Basic XPath: {item['XPath (ID)']}")
        print(f"   Relative XPath: {item['Relative XPath']}")
        print(f"   Absolute XPath: {item['Absolute XPath']}")
        print()
    
    print(f"   ... and {len(data) - 2} more elements with full locator mappings")
    print("=" * 100)
    print(f"\n🎯 Ready for comprehensive automation testing with multiple XPath strategies!")
    
    return df

if __name__ == "__main__":
    result_df = main()