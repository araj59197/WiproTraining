from openpyxl import load_workbook, Workbook

# Load the Excel file
wb = load_workbook('sales_data.xlsx')
ws = wb['2025']

# Add "Total" header in column D
ws['D1'] = 'Total'

# Calculate and add Total values for each row
for row in range(2, ws.max_row + 1):
    quantity = ws[f'B{row}'].value
    price = ws[f'C{row}'].value
    total = quantity * price
    ws[f'D{row}'] = total

# Display the data
print("Updated Data:")
for row in ws.iter_rows(min_row=1, values_only=True):
    print(row)

# Save to a new Excel file
wb.save('sales_summary_openpyxl.xlsx')

print("\nData successfully saved to sales_summary_openpyxl.xlsx")
